"""Reads a list of job posting URLs from a .csv or .xlsx file for batch
tailoring (see cli.py's `batch` command).

Requires a header row with a recognizable URL column name rather than
falling back to "just use the first column" when nothing matches — a
silent wrong guess would happily scrape a header string as if it were a
job posting URL, burning a real API call before failing far downstream
with a confusing error. Failing fast, at file-read time, with a clear
message is cheaper and clearer for the user than that.
"""

from pathlib import Path

_URL_COLUMN_CANDIDATES = ["job_url", "url", "link", "job_link"]


class BatchFileError(ValueError):
    pass


def _normalize_header(header: str) -> str:
    """Lowercases and collapses spaces/hyphens to underscores, so a
    naturally-written header like "Job URL" or "job-url" still matches
    the "job_url" candidate — not just an exact underscore-separated
    match."""
    return "_".join(header.strip().lower().split()).replace("-", "_")


def _pick_column(headers: list[str]) -> int:
    normalized = [_normalize_header(h) for h in headers]
    for candidate in _URL_COLUMN_CANDIDATES:
        if candidate in normalized:
            return normalized.index(candidate)
    raise BatchFileError(
        f"Couldn't find a URL column in the header row — expected one of "
        f"{_URL_COLUMN_CANDIDATES}, found: {headers}"
    )


def _read_csv(path: Path) -> list[str]:
    import csv

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = [row for row in csv.reader(f) if any(cell.strip() for cell in row)]

    if not rows:
        return []
    header, *data_rows = rows
    col = _pick_column(header)
    return [row[col].strip() for row in data_rows if col < len(row) and row[col].strip()]


def _read_xlsx(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows = [
            row for row in sheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)
        ]
    finally:
        wb.close()

    if not rows:
        return []
    header, *data_rows = rows
    header = [str(h) if h is not None else "" for h in header]
    col = _pick_column(header)
    return [
        str(row[col]).strip()
        for row in data_rows
        if col < len(row) and row[col] not in (None, "")
    ]


def read_job_urls(path: str) -> list[str]:
    """Read job posting URLs from a .csv or .xlsx file. Expects a header
    row with a job_url/url/link/job_link column (case-insensitive).
    Blank rows and blank cells in that column are skipped. Duplicate URLs
    are de-duplicated (first occurrence wins, order preserved) — running
    the same posting twice in one batch just wastes API calls for an
    identical result.
    """
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        urls = _read_csv(p)
    elif suffix in (".xlsx", ".xlsm"):
        urls = _read_xlsx(p)
    else:
        raise BatchFileError(f"Unsupported batch file type '{suffix}' — use .csv or .xlsx")

    urls = list(dict.fromkeys(urls))
    if not urls:
        raise BatchFileError(f"No job URLs found in {path}")
    return urls
